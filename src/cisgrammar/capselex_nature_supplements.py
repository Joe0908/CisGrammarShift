from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from cisgrammar.capselex import normalize_tf_symbol
from cisgrammar.capselex_monomer_motifs import PWM

DNA = frozenset("ACGT")


def _normal_pair(first: object, second: object) -> tuple[str, str]:
    return normalize_tf_symbol(first), normalize_tf_symbol(second)


def _pair_members(pair: str) -> tuple[str, str] | None:
    parts = pair.split("_", 1)
    if len(parts) != 2:
        return None
    return _normal_pair(parts[0], parts[1])


@dataclass(frozen=True)
class InteractionRecord:
    raw_bait: str
    raw_prey: str
    bait: str
    prey: str
    raw_value: str
    tested: bool
    composite_motif: bool
    spacing_or_orientation: bool
    specificity: tuple[str, ...]

    @property
    def raw_pair(self) -> str:
        return f"{self.raw_bait}_{self.raw_prey}"

    @property
    def pair(self) -> str:
        return f"{self.bait}_{self.prey}"

    @property
    def positive(self) -> bool:
        return self.composite_motif or self.spacing_or_orientation


@dataclass(frozen=True)
class CapPWMRecord:
    raw_pair: str
    pair: str
    ligand_sequence: str
    batch: str
    seed: str
    multinomial: int
    cycle: str
    interaction_type: str | None
    representative: bool | None
    counts: np.ndarray
    source_row: int

    @property
    def members(self) -> tuple[str, str] | None:
        return _pair_members(self.pair)

    @property
    def model_id(self) -> str:
        fields = (
            self.pair,
            self.ligand_sequence,
            self.batch,
            self.seed,
            str(self.multinomial),
            self.cycle,
        )
        return "|".join(fields)

    def to_pwm(self) -> PWM:
        return PWM(self.model_id, self.counts)


@dataclass(frozen=True)
class CapSpacingRecord:
    raw_pair: str
    pair: str
    barcode: str
    batch: str
    analysis_type: str
    left_kmer: str
    right_kmer: str
    gap_counts: tuple[int, ...]
    source_row: int

    @property
    def members(self) -> tuple[str, str]:
        parsed = _pair_members(self.pair)
        if parsed is None:
            raise ValueError(f"spacing record is not a TF pair: {self.pair}")
        return parsed


def _parse_interaction_value(value: object) -> tuple[bool, bool, bool, tuple[str, ...]]:
    raw = str(value).strip()
    if raw == "0":
        return True, False, False, ()
    if raw.upper() == "N/A" or raw.lower() == "self":
        return False, False, False, ()
    tokens = tuple(part.strip().lower() for part in raw.split(",") if part.strip())
    composite = any(token.startswith("1-") for token in tokens)
    spacing = any(token.startswith("2-") for token in tokens)
    if not composite and not spacing:
        raise ValueError(f"unrecognized interaction value: {value!r}")
    specificity = tuple(sorted({token.split("-", 1)[1] for token in tokens}))
    return True, composite, spacing, specificity


def read_interaction_matrix(path: str | Path) -> list[InteractionRecord]:
    """Read the directed bait-column/prey-row matrix in Nature Supplementary Table 2."""
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows = worksheet.iter_rows(values_only=True)
    for row in rows:
        if row and row[0] == "Prey\\Bait":
            baits = [str(value).strip() for value in row[1:] if value is not None]
            break
    else:
        raise ValueError("Supplementary Table 2 interaction header was not found")
    if not baits or len(set(baits)) != len(baits):
        raise ValueError("interaction bait names must be present and unique")

    records: list[InteractionRecord] = []
    for row in rows:
        prey_value = row[0] if row else None
        if isinstance(prey_value, str) and prey_value.startswith("Section "):
            break
        if prey_value is None:
            continue
        raw_prey = str(prey_value).strip()
        for raw_bait, value in zip(baits, row[1 : len(baits) + 1], strict=True):
            if value is None:
                continue
            tested, composite, spacing, specificity = _parse_interaction_value(value)
            bait, prey = _normal_pair(raw_bait, raw_prey)
            records.append(
                InteractionRecord(
                    raw_bait=raw_bait,
                    raw_prey=raw_prey,
                    bait=bait,
                    prey=prey,
                    raw_value=str(value).strip(),
                    tested=tested,
                    composite_motif=composite,
                    spacing_or_orientation=spacing,
                    specificity=specificity,
                )
            )
    if not records:
        raise ValueError("Supplementary Table 2 interaction matrix is empty")
    return records


def _as_optional_bool(value: object) -> bool | None:
    if value is None:
        return None
    normalized = str(value).strip().upper()
    if normalized == "YES":
        return True
    if normalized == "NO":
        return False
    raise ValueError(f"representative PWM flag must be YES, NO or empty; got {value!r}")


def read_pwm_models(path: str | Path) -> list[CapPWMRecord]:
    """Read count matrices and metadata from Nature Supplementary Table 3."""
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows = iter(enumerate(worksheet.iter_rows(values_only=True), start=1))
    records: list[CapPWMRecord] = []
    for source_row, row in rows:
        if (
            not row
            or row[0] != "Base"
            or row[1] in {None, "TF_pairs"}
            or len(row) < 7
            or row[2] is None
        ):
            continue
        raw_pair = str(row[1]).strip()
        pair_members = _pair_members(raw_pair)
        pair = "_".join(pair_members) if pair_members is not None else normalize_tf_symbol(raw_pair)
        base_counts: dict[str, list[float]] = {}
        for expected_base in "ACGT":
            try:
                _, base_row = next(rows)
            except StopIteration as error:
                raise ValueError(f"incomplete PWM beginning at row {source_row}") from error
            if base_row[0] != expected_base:
                raise ValueError(
                    f"PWM beginning at row {source_row} expected {expected_base}, got {base_row[0]!r}"
                )
            base_counts[expected_base] = [float(value) for value in base_row[1:] if value is not None]
        lengths = {len(values) for values in base_counts.values()}
        if len(lengths) != 1 or not lengths or next(iter(lengths)) == 0:
            raise ValueError(f"PWM beginning at row {source_row} has inconsistent base lengths")
        counts = np.array([base_counts[base] for base in "ACGT"], dtype=float).T
        if not np.isfinite(counts).all() or np.any(counts < 0):
            raise ValueError(f"PWM beginning at row {source_row} has invalid counts")
        interaction_type = str(row[7]).strip().lower() if row[7] is not None else None
        if interaction_type not in {None, "composite", "spacing"}:
            raise ValueError(f"unknown PWM interaction type at row {source_row}: {row[7]!r}")
        records.append(
            CapPWMRecord(
                raw_pair=raw_pair,
                pair=pair,
                ligand_sequence=str(row[2]).strip(),
                batch=str(row[3]).strip(),
                seed=str(row[4]).strip(),
                multinomial=int(row[5]),
                cycle=str(row[6]).strip(),
                interaction_type=interaction_type,
                representative=_as_optional_bool(row[8]),
                counts=counts,
                source_row=source_row,
            )
        )
    if not records:
        raise ValueError("Supplementary Table 3 contains no PWM models")
    model_ids = [record.model_id for record in records]
    if len(set(model_ids)) != len(model_ids):
        raise ValueError("Supplementary Table 3 PWM model identifiers are not unique")
    return records


def read_spacing_counts(path: str | Path) -> list[CapSpacingRecord]:
    """Read oriented 6-mer counts across gaps 0--28 from Supplementary Table 7."""
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows = worksheet.iter_rows(values_only=True)
    data_start_row: int | None = None
    for header_row, row in enumerate(rows, start=1):
        if row and row[0] == "TF1_TF2" and row[4] == "kmer1.kmer2":
            gap_headers = row[5:]
            expected = tuple(f"Count Gap {gap}" for gap in range(29))
            observed = tuple(str(value) for value in gap_headers if value is not None)
            if observed != expected:
                raise ValueError("Supplementary Table 7 must contain Count Gap 0 through Count Gap 28")
            data_start_row = header_row + 1
            break
    else:
        raise ValueError("Supplementary Table 7 spacing header was not found")

    if data_start_row is None:
        raise AssertionError("spacing data row was not resolved")
    records: list[CapSpacingRecord] = []
    for source_row, row in enumerate(rows, start=data_start_row):
        if not row or row[0] is None:
            continue
        raw_pair = str(row[0]).strip()
        members = _pair_members(raw_pair)
        if members is None:
            raise ValueError(f"invalid TF pair at row {source_row}: {raw_pair!r}")
        kmers = str(row[4]).upper().split(".", 1)
        if len(kmers) != 2 or any(set(kmer) - DNA for kmer in kmers):
            raise ValueError(f"invalid oriented k-mer pair at row {source_row}: {row[4]!r}")
        counts = tuple(int(value) for value in row[5:34])
        if len(counts) != 29 or any(value < 0 for value in counts):
            raise ValueError(f"invalid gap counts at row {source_row}")
        records.append(
            CapSpacingRecord(
                raw_pair=raw_pair,
                pair="_".join(members),
                barcode=str(row[1]).strip(),
                batch=str(row[2]).strip(),
                analysis_type=str(row[3]).strip(),
                left_kmer=kmers[0],
                right_kmer=kmers[1],
                gap_counts=counts,
                source_row=source_row,
            )
        )
    if not records:
        raise ValueError("Supplementary Table 7 contains no spacing records")
    return records


def _pairs_with_tf(pairs: set[str], tf: str) -> set[str]:
    normalized = normalize_tf_symbol(tf)
    return {pair for pair in pairs if (members := _pair_members(pair)) and normalized in members}


def audit_supplement_coverage(
    interactions: list[InteractionRecord],
    pwm_models: list[CapPWMRecord],
    spacing_counts: list[CapSpacingRecord],
    focal_tfs: list[str] | tuple[str, ...],
) -> dict[str, object]:
    composite_interactions = {record.pair for record in interactions if record.composite_motif}
    spacing_interactions = {record.pair for record in interactions if record.spacing_or_orientation}
    composite_pwm_pairs = {
        record.pair for record in pwm_models if record.interaction_type == "composite"
    }
    spacing_pwm_pairs = {record.pair for record in pwm_models if record.interaction_type == "spacing"}
    spacing_count_pairs = {record.pair for record in spacing_counts}

    panel: list[dict[str, object]] = []
    for focal_tf in focal_tfs:
        tf = normalize_tf_symbol(focal_tf)
        focal_models = [
            record for record in pwm_models if record.members is not None and tf in record.members
        ]
        focal_spacing_rows = [record for record in spacing_counts if tf in record.members]
        representative_models = [record for record in focal_models if record.representative]
        positive_pairs = _pairs_with_tf(composite_interactions | spacing_interactions, tf)
        partners = sorted(
            {
                member
                for pair in positive_pairs
                for member in (_pair_members(pair) or ())
                if member != tf
            }
        )
        panel.append(
            {
                "tf": tf,
                "positive_directed_pairs": len(positive_pairs),
                "positive_partners": partners,
                "composite_positive_pairs": len(_pairs_with_tf(composite_interactions, tf)),
                "spacing_positive_pairs": len(_pairs_with_tf(spacing_interactions, tf)),
                "pwm_models": len(focal_models),
                "pwm_pairs": len({record.pair for record in focal_models}),
                "composite_pwm_models": sum(
                    record.interaction_type == "composite" for record in focal_models
                ),
                "spacing_pwm_models": sum(record.interaction_type == "spacing" for record in focal_models),
                "representative_pwm_models": len(representative_models),
                "representative_pwm_pairs": sorted({record.pair for record in representative_models}),
                "spacing_count_rows": len(focal_spacing_rows),
                "spacing_count_pairs": len({record.pair for record in focal_spacing_rows}),
                "raw_cap_feature_assets_available": bool(focal_models or focal_spacing_rows),
                "primary_eligible": None,
                "primary_eligibility_note": (
                    "Raw CAP coverage only; partner expression, monomer controls and the frozen "
                    "feature-selection rule must be checked separately."
                ),
            }
        )

    interaction_values = Counter(record.raw_value for record in interactions)
    return {
        "schema_version": "capselex_nature_supplement_audit_v1",
        "global": {
            "interaction_matrix_records": len(interactions),
            "tested_directed_cells": sum(record.tested for record in interactions),
            "positive_directed_cells": sum(record.positive for record in interactions),
            "composite_positive_pairs": len(composite_interactions),
            "spacing_positive_pairs": len(spacing_interactions),
            "positive_pair_union": len(composite_interactions | spacing_interactions),
            "counting_note": (
                "Literal directed cells in Supplementary Table 2; do not substitute this count for "
                "the article's separately deduplicated headline total."
            ),
            "interaction_value_counts": dict(sorted(interaction_values.items())),
            "pwm_models": len(pwm_models),
            "pwm_pair_models": sum(record.members is not None for record in pwm_models),
            "individual_pwm_models": sum(record.members is None for record in pwm_models),
            "composite_pwm_models": sum(
                record.interaction_type == "composite" for record in pwm_models
            ),
            "spacing_pwm_models": sum(record.interaction_type == "spacing" for record in pwm_models),
            "representative_pwm_models": sum(record.representative is True for record in pwm_models),
            "spacing_count_rows": len(spacing_counts),
            "spacing_count_pairs": len(spacing_count_pairs),
        },
        "cross_table_consistency": {
            "composite_interactions_without_pwm_pair": sorted(
                composite_interactions - composite_pwm_pairs
            ),
            "composite_pwm_pairs_without_interaction": sorted(
                composite_pwm_pairs - composite_interactions
            ),
            "spacing_interactions_without_count_pair": sorted(
                spacing_interactions - spacing_count_pairs
            ),
            "spacing_count_pairs_without_interaction": sorted(
                spacing_count_pairs - spacing_interactions
            ),
            "spacing_pwm_pairs_not_in_spacing_interactions": sorted(
                spacing_pwm_pairs - spacing_interactions
            ),
            "alias_normalization_applied": True,
        },
        "panel": panel,
        "claim_boundary": (
            "This audit establishes public-file availability and cross-table coverage. It does not "
            "establish a CAP effect on ChIP, a present partner, or causal cooperativity."
        ),
    }
