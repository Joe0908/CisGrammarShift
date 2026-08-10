from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from cisgrammar.capselex import asset_manifest, normalize_tf_symbol

_EXPERIMENT_CYCLE = re.compile(r"\bExp\.\s+(GHT\d+)_(\d+)\b", re.IGNORECASE)


@dataclass(frozen=True)
class GHTSelectedRun:
    tf: str
    run_accession: str
    experiment_id: str
    cycle: int
    peak_filename: str


@dataclass(frozen=True)
class ENAFastqFile:
    url: str
    md5: str
    size_bytes: int


@dataclass(frozen=True)
class GHTExperiment:
    experiment_id: str
    tf: str
    batch_key: str
    approved: bool


def _metadata_sheet_and_indices(metadata_path: str | Path):
    workbook = load_workbook(metadata_path, read_only=True, data_only=True)
    if "Metadata" not in workbook.sheetnames:
        raise ValueError("GHT workbook has no Metadata sheet")
    sheet = workbook["Metadata"]
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        headers = [str(value).strip() if value is not None else "" for value in row]
        if "*title" in headers and "*SRA Experiment or Run" in headers:
            title_index = headers.index("*title")
            run_indices = [
                index for index, value in enumerate(headers) if value == "*SRA Experiment or Run"
            ]
            processed_indices = [
                index for index, value in enumerate(headers) if value.strip() == "processed data file"
            ]
            if not run_indices or not processed_indices:
                raise ValueError("GHT workbook lacks SRA run or processed-data columns")
            return sheet, row_number, title_index, run_indices, processed_indices[-1]
    raise ValueError("GHT workbook metadata header was not found")


def read_ght_run_inventory(metadata_path: str | Path) -> dict[str, list[tuple[int, str]]]:
    sheet, header_number, title_index, run_indices, _ = _metadata_sheet_and_indices(metadata_path)
    by_cycle: dict[str, dict[int, str]] = {}
    for row in sheet.iter_rows(min_row=header_number + 1, values_only=True):
        run_accession = next(
            (str(row[index]).strip() for index in run_indices if row[index]),
            "",
        )
        if not run_accession:
            continue
        title = str(row[title_index] or "")
        match = _EXPERIMENT_CYCLE.search(title)
        if match is None:
            continue
        experiment_id = match.group(1).upper()
        cycle = int(match.group(2))
        existing = by_cycle.setdefault(experiment_id, {}).get(cycle)
        if existing is not None and existing != run_accession:
            raise ValueError(
                f"conflicting runs for {experiment_id} cycle {cycle}: {existing}, {run_accession}"
            )
        by_cycle[experiment_id][cycle] = run_accession
    return {
        experiment_id: sorted(records.items()) for experiment_id, records in by_cycle.items()
    }


def _batch_key(batch_id: object, greco_id: object) -> str:
    batch = str(batch_id or "").strip()
    greco = str(greco_id or "").strip()
    if not batch or not greco:
        raise ValueError("GHT experiment lacks a batch or Greco experiment ID")
    suffix = greco.removeprefix(batch + "_").split("_", 1)[0]
    return f"{batch}_{suffix}" if len(suffix) == 1 and suffix.isalpha() else batch


def read_ght_experiments(path: str | Path) -> list[GHTExperiment]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "TableS3" not in workbook.sheetnames:
        raise ValueError("GHT supplementary metadata has no TableS3 sheet")
    rows = workbook["TableS3"].iter_rows(values_only=True)
    headers = next(rows)
    indices = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
    required = {"UID", "Greco Experiment ID", "TF Id", "Batch ID", "Approved", "Assay"}
    missing = required - set(indices)
    if missing:
        raise ValueError(f"GHT TableS3 is missing columns: {', '.join(sorted(missing))}")
    experiments = []
    for row in rows:
        if str(row[indices["Assay"]] or "").strip() != "GHT-SELEX":
            continue
        experiment_id = str(row[indices["UID"]] or "").strip().upper()
        if not experiment_id:
            continue
        approved_value = row[indices["Approved"]]
        approved = approved_value is True or str(approved_value).strip().lower() == "true"
        experiments.append(
            GHTExperiment(
                experiment_id=experiment_id,
                tf=normalize_tf_symbol(row[indices["TF Id"]]),
                batch_key=_batch_key(
                    row[indices["Batch ID"]],
                    row[indices["Greco Experiment ID"]],
                ),
                approved=approved,
            )
        )
    duplicates = [
        experiment_id
        for experiment_id in {record.experiment_id for record in experiments}
        if sum(record.experiment_id == experiment_id for record in experiments) > 1
    ]
    if duplicates:
        raise ValueError(f"duplicate GHT experiment IDs in TableS3: {', '.join(sorted(duplicates))}")
    return experiments


def read_selected_ght_runs(
    metadata_path: str | Path,
    focal_tfs: list[str],
) -> list[GHTSelectedRun]:
    """Read runs contributing to author-provided MAGIX peak sets.

    The corrected GEO workbook has duplicate SRA and processed-file headers and leaves
    several descriptive columns empty. The final processed-file column is therefore used
    as the authoritative link between a run and an author-selected merged MAGIX peak set.
    Runs without that filename were not included in the published merged peak set.
    """

    requested = {normalize_tf_symbol(tf) for tf in focal_tfs}
    sheet, header_number, title_index, run_indices, peak_index = _metadata_sheet_and_indices(
        metadata_path
    )

    selected: list[GHTSelectedRun] = []
    for row in sheet.iter_rows(min_row=header_number + 1, values_only=True):
        peak_filename = str(row[peak_index] or "").strip()
        if not peak_filename:
            continue
        tf = normalize_tf_symbol(peak_filename.split("_MAGIX_", 1)[0])
        if tf not in requested or "_MAGIX_" not in peak_filename:
            continue
        run_accession = next(
            (str(row[index]).strip() for index in run_indices if row[index]),
            "",
        )
        if not run_accession:
            raise ValueError(f"{peak_filename} has a selected row without an SRA run accession")
        title = str(row[title_index] or "")
        match = _EXPERIMENT_CYCLE.search(title)
        if match is None:
            raise ValueError(f"cannot parse experiment/cycle from title: {title}")
        selected.append(
            GHTSelectedRun(
                tf=tf,
                run_accession=run_accession,
                experiment_id=match.group(1).upper(),
                cycle=int(match.group(2)),
                peak_filename=peak_filename,
            )
        )

    duplicates = [
        accession
        for accession in {record.run_accession for record in selected}
        if sum(record.run_accession == accession for record in selected) > 1
    ]
    if duplicates:
        raise ValueError(f"duplicate selected SRA runs: {', '.join(sorted(duplicates))}")
    missing = sorted(requested - {record.tf for record in selected})
    if missing:
        raise ValueError(f"no author-selected GHT runs for: {', '.join(missing)}")
    return sorted(selected, key=lambda record: (record.tf, record.experiment_id, record.cycle))


def read_ena_fastq_report(path: str | Path) -> dict[str, list[ENAFastqFile]]:
    required = {"run_accession", "fastq_ftp", "fastq_md5", "fastq_bytes"}
    result: dict[str, list[ENAFastqFile]] = {}
    with Path(path).open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"ENA report is missing columns: {', '.join(sorted(missing))}")
        for row in reader:
            accession = row["run_accession"].strip()
            urls = [value.strip() for value in row["fastq_ftp"].split(";") if value.strip()]
            md5s = [value.strip() for value in row["fastq_md5"].split(";") if value.strip()]
            sizes = [value.strip() for value in row["fastq_bytes"].split(";") if value.strip()]
            if not urls:
                continue
            if not (len(urls) == len(md5s) == len(sizes)):
                raise ValueError(f"ENA FASTQ fields have unequal lengths for {accession}")
            result[accession] = [
                ENAFastqFile(
                    url=url if "://" in url else f"https://{url}",
                    md5=md5,
                    size_bytes=int(size),
                )
                for url, md5, size in zip(urls, md5s, sizes, strict=True)
            ]
    return result


def build_ght_rebuild_audit(
    metadata_path: str | Path,
    experiment_metadata_path: str | Path,
    ena_report_path: str | Path,
    primary_tfs: list[str],
    sensitivity_tfs: list[str],
    audit_date: str,
    magix_commit: str,
) -> dict[str, object]:
    primary = [normalize_tf_symbol(tf) for tf in primary_tfs]
    sensitivity = [normalize_tf_symbol(tf) for tf in sensitivity_tfs]
    if set(primary) & set(sensitivity):
        raise ValueError("primary and sensitivity TF panels must be disjoint")
    runs = read_selected_ght_runs(metadata_path, primary + sensitivity)
    run_inventory = read_ght_run_inventory(metadata_path)
    experiments = read_ght_experiments(experiment_metadata_path)
    experiments_by_id = {record.experiment_id: record for record in experiments}
    ena = read_ena_fastq_report(ena_report_path)
    missing_runs = sorted({record.run_accession for record in runs} - set(ena))
    if missing_runs:
        raise ValueError(f"ENA report lacks selected runs: {', '.join(missing_runs)}")
    missing_experiments = sorted({record.experiment_id for record in runs} - set(experiments_by_id))
    if missing_experiments:
        raise ValueError(
            f"TableS3 lacks selected experiments: {', '.join(missing_experiments)}"
        )
    for record in runs:
        experiment = experiments_by_id[record.experiment_id]
        if experiment.tf != record.tf or not experiment.approved:
            raise ValueError(f"selected run disagrees with approved TableS3 metadata: {record.run_accession}")

    panel = []
    for tf in primary + sensitivity:
        tf_runs = [record for record in runs if record.tf == tf]
        batch_keys = sorted({experiments_by_id[record.experiment_id].batch_key for record in tf_runs})
        serialized_runs = []
        for record in tf_runs:
            files = ena[record.run_accession]
            serialized_runs.append(
                {
                    "run_accession": record.run_accession,
                    "experiment_id": record.experiment_id,
                    "cycle": record.cycle,
                    "read_layout": "paired" if len(files) == 2 else "single",
                    "fastq_files": [
                        {"url": item.url, "md5": item.md5, "size_bytes": item.size_bytes}
                        for item in files
                    ],
                    "total_size_bytes": sum(item.size_bytes for item in files),
                }
            )
        panel.append(
            {
                "tf": tf,
                "role": "primary" if tf in primary else "low_cap_coverage_sensitivity",
                "magix_peak_filename": tf_runs[0].peak_filename,
                "batch_keys": batch_keys,
                "run_count": len(tf_runs),
                "fastq_file_count": sum(len(record["fastq_files"]) for record in serialized_runs),
                "total_size_bytes": sum(record["total_size_bytes"] for record in serialized_runs),
                "runs": serialized_runs,
            }
        )

    primary_bytes = sum(record["total_size_bytes"] for record in panel if record["role"] == "primary")
    all_bytes = sum(record["total_size_bytes"] for record in panel)

    def batch_scope(batch_keys: set[str], approved_only: bool) -> dict[str, object]:
        scoped_experiments = [
            record
            for record in experiments
            if record.batch_key in batch_keys and (record.approved or not approved_only)
        ]
        scoped_runs = sorted(
            {
                run_accession
                for experiment in scoped_experiments
                for _, run_accession in run_inventory.get(experiment.experiment_id, [])
            }
        )
        missing = sorted(set(scoped_runs) - set(ena))
        if missing:
            raise ValueError(f"ENA report lacks batch-background runs: {', '.join(missing)}")
        files = [item for accession in scoped_runs for item in ena[accession]]
        return {
            "experiment_count": len(scoped_experiments),
            "experiments_with_geo_runs": sum(
                experiment.experiment_id in run_inventory for experiment in scoped_experiments
            ),
            "run_count": len(scoped_runs),
            "fastq_file_count": len(files),
            "total_size_bytes": sum(item.size_bytes for item in files),
            "total_size_gib": round(sum(item.size_bytes for item in files) / 2**30, 3),
        }

    primary_batch_keys = {
        batch_key
        for record in panel
        if record["role"] == "primary"
        for batch_key in record["batch_keys"]
    }
    all_batch_keys = {batch_key for record in panel for batch_key in record["batch_keys"]}
    batch_scenarios = {
        "primary_panel": {
            "batch_keys": sorted(primary_batch_keys),
            "all_ght_experiments_in_batches": batch_scope(primary_batch_keys, False),
            "approved_ght_experiments_only": batch_scope(primary_batch_keys, True),
        },
        "all_panel_including_sensitivity": {
            "batch_keys": sorted(all_batch_keys),
            "all_ght_experiments_in_batches": batch_scope(all_batch_keys, False),
            "approved_ght_experiments_only": batch_scope(all_batch_keys, True),
        },
    }
    return {
        "schema_version": "codebook_ght_v2_rebuild_audit_v1",
        "audit_date": audit_date,
        "source_assets": asset_manifest(
            [Path(metadata_path), Path(experiment_metadata_path), Path(ena_report_path)]
        ),
        "magix_source": {
            "repository": "https://github.com/csglab/MAGIX",
            "commit": magix_commit,
            "release_doi": "10.5281/zenodo.20846978",
            "license": "GPL-3.0",
        },
        "selection_rule": (
            "Include only runs whose final processed-data column names the author-provided merged "
            "MAGIX peak set for the focal TF. Rows without that filename are not silently added."
        ),
        "primary_target_fastq_bytes": primary_bytes,
        "all_target_fastq_bytes": all_bytes,
        "primary_target_fastq_gib": round(primary_bytes / 2**30, 3),
        "all_target_fastq_gib": round(all_bytes / 2**30, 3),
        "target_fastq_download_ready": True,
        "exact_author_v2_rebuild_ready": False,
        "batch_background_download_scenarios": batch_scenarios,
        "exact_rebuild_blockers": [
            "The MAGIX production design matrices and batch file lists are not tracked in the public "
            "source repository. Supplementary Table 3 exposes batch membership, but the released "
            "materials do not state whether production aggregates include every experiment or only "
            "approved experiments.",
            "The published workflow fits batch-aggregate covariates; focal target FASTQs alone do not "
            "reproduce those aggregate counts or the authors' fitted library-size model. The two "
            "plausible aggregate scopes have materially different acquisition costs and cannot be "
            "chosen after inspecting outcomes.",
            "The corrected v2 MAGIX BED archive was unavailable from the Codebook host on the audit date.",
        ],
        "resource_note": (
            "The public production script requests 100 GB RAM for the approximately 13 million-bin "
            "genome-wide library-size fit and 20 GB RAM for focal candidate-peak fitting."
        ),
        "panel": panel,
        "claim_boundary": (
            "This audit proves that focal target reads are small enough to acquire selectively. It does "
            "not claim exact reproduction of corrected v2 MAGIX peaks until the production design and "
            "batch aggregate membership rule, or the official corrected BED files, are available."
        ),
    }
