from __future__ import annotations

import numpy as np
from openpyxl import Workbook

from cisgrammar.capselex_nature_supplements import (
    audit_supplement_coverage,
    read_interaction_matrix,
    read_pwm_models,
    read_spacing_counts,
)


def _save_interactions(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Section a"])
    sheet.append(["Prey\\Bait", "FLI1", "RFXDC2"])
    sheet.append(["GCM1", "1-specific,2-subfamily", 0])
    sheet.append(["OTP", "N/A", "2-specific"])
    sheet.append(["Section b"])
    workbook.save(path)


def _save_pwms(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None])
    sheet.append([None, "TF_pairs"])
    sheet.append(["Base", "FLI1_GCM1", "barcode", "batch", "NACGTN", 1, 3, "composite", "YES"])
    for base, values in zip("ACGT", np.eye(4, dtype=int), strict=True):
        sheet.append([base, *values])
    workbook.save(path)


def _save_spacing(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None])
    sheet.append(
        ["TF1_TF2", "Barcode", "Batch", None, "kmer1.kmer2"]
        + [f"Count Gap {gap}" for gap in range(29)]
    )
    sheet.append(["FLI1_GCM1", "barcode", "batch", "Halfsite_spacing", "AAAAAA.CCCCCC"] + list(range(29)))
    sheet.append(["RFX7_OTP", "barcode", "batch", "Halfsite_spacing", "AAAAAA.CCCCCC"] + list(range(29)))
    workbook.save(path)


def test_official_supplement_parsers_and_alias_consistency(tmp_path) -> None:
    interactions_path = tmp_path / "interactions.xlsx"
    pwm_path = tmp_path / "pwms.xlsx"
    spacing_path = tmp_path / "spacing.xlsx"
    _save_interactions(interactions_path)
    _save_pwms(pwm_path)
    _save_spacing(spacing_path)

    interactions = read_interaction_matrix(interactions_path)
    pwm_models = read_pwm_models(pwm_path)
    spacing = read_spacing_counts(spacing_path)
    report = audit_supplement_coverage(interactions, pwm_models, spacing, ["FLI1", "RFX7"])

    assert len(interactions) == 4
    assert interactions[0].pair == "FLI1_GCM1"
    assert interactions[0].composite_motif and interactions[0].spacing_or_orientation
    assert pwm_models[0].counts.shape == (4, 4)
    assert pwm_models[0].to_pwm().length == 4
    assert spacing[1].pair == "RFX7_OTP"
    assert spacing[1].gap_counts[-1] == 28
    assert report["cross_table_consistency"]["spacing_interactions_without_count_pair"] == []
    assert report["panel"][0]["representative_pwm_models"] == 1


def test_pwm_parser_rejects_incomplete_base_block(tmp_path) -> None:
    path = tmp_path / "bad.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Base", "FLI1_GCM1", "barcode", "batch", "NACGTN", 1, 3, "composite", "YES"])
    sheet.append(["A", 1, 2])
    workbook.save(path)

    try:
        read_pwm_models(path)
    except ValueError as error:
        assert "incomplete PWM" in str(error)
    else:
        raise AssertionError("incomplete PWM should fail")
